"""One-time seed import for the DR SMB Outreach CRM.

Reads research/san_cristobal_candidates.xlsx (Candidates tab) — 42 researched
companies — and inserts each as an 'Identified' company in San Cristóbal.
Optionally imports any real rows from tracker/DR_Outreach_Tracker.xlsx (Companies
tab), skipping the template/example row. The reference-only tabs
("Excluded (Already Digital)", "Hospitals (Not Recommended Now)") are ignored.

Usage:
    python seed_import.py            # seed into an empty DB
    python seed_import.py --reset    # wipe existing companies, then reseed

Run from the crm/ directory (or anywhere; paths are resolved relative to the
project root, one level up from this file).
"""
import os
import sys

# Windows consoles default to cp1252 and choke on the accented company names /
# box-drawing chars in the summary. Force UTF-8 output.
try:
    sys.stdout.reconfigure(encoding="utf-8")
except (AttributeError, ValueError):
    pass

import openpyxl

import db
from domain import SECTORS, SIZES, SOURCES

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(BASE_DIR)


def _find(*candidates):
    """Return the first path that exists (or the last, for a clear error)."""
    for p in candidates:
        if os.path.exists(p):
            return p
    return candidates[-1]


# Prefer the original project files (local dev); fall back to the copies bundled
# in crm/seed_data/ so the app is self-contained when deployed on its own.
CANDIDATES_XLSX = _find(
    os.path.join(PROJECT_ROOT, "research", "san_cristobal_candidates.xlsx"),
    os.path.join(BASE_DIR, "seed_data", "san_cristobal_candidates.xlsx"),
)
TRACKER_XLSX = _find(
    os.path.join(PROJECT_ROOT, "tracker", "DR_Outreach_Tracker.xlsx"),
    os.path.join(BASE_DIR, "seed_data", "DR_Outreach_Tracker.xlsx"),
)

# The example/demo row shipped in the tracker template — never import it.
TRACKER_EXAMPLE_NAME = "Ferretería Pérez"


def _clean(v):
    if v is None:
        return None
    v = str(v).strip()
    return v or None


def map_size(raw):
    """Enum allows only Small/Medium/Large. Research uses 'Small-Medium' too."""
    raw = _clean(raw)
    if raw in SIZES:
        return raw
    if raw == "Small-Medium":
        return "Medium"  # lean up; original preserved in notes
    return None


def map_source(raw):
    """Map research sources onto the Source enum; directories become 'Otro'."""
    raw = _clean(raw)
    if raw in SOURCES:
        return raw
    return "Otro"  # e.g. 'Yelu.do directory', 'Directory search'


def map_sector(raw):
    raw = _clean(raw)
    return raw if raw in SECTORS else "Otro"


def build_notes(confidence, orig_source, social, orig_notes, orig_size=None):
    """Preserve the research context that has no dedicated column."""
    parts = []
    if confidence:
        parts.append(f"Confidence: {confidence}.")
    if orig_source:
        parts.append(f"Source: {orig_source}.")
    if social:
        parts.append(f"Social: {social}.")
    if orig_size and orig_size not in SIZES:
        parts.append(f"Est. size (raw): {orig_size}.")
    if orig_notes:
        parts.append(str(orig_notes))
    return " ".join(parts).strip() or None


def import_candidates(conn):
    wb = openpyxl.load_workbook(CANDIDATES_XLSX, data_only=True)
    ws = wb["Candidates"]
    # Columns: Company Name, Sector, Neighborhood/City, Phone, Social Media
    # Handle, Estimated Size, Confidence (No Website), Source, Notes/Signal
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = _clean(row[0])
        if not name:
            continue
        social = _clean(row[4])
        data = {
            "company_name": name,
            "sector": map_sector(row[1]),
            "province_city": "San Cristóbal",
            "address": _clean(row[2]),                 # Neighborhood/City
            "phone": _clean(row[3]),
            "has_website": "N",                         # this is the no-website list
            "has_social_only": "Y" if social else "N",
            "estimated_size": map_size(row[5]),
            "source": map_source(row[7]),
            "current_stage": "Identified",
            "notes": build_notes(_clean(row[6]), _clean(row[7]), social,
                                  _clean(row[8]), _clean(row[5])),
        }
        db.insert_company(conn, data)
        count += 1
    return count


def import_tracker(conn):
    """Import real rows from the tracker Companies tab, skipping the example."""
    if not os.path.exists(TRACKER_XLSX):
        return 0
    wb = openpyxl.load_workbook(TRACKER_XLSX, data_only=True)
    if "Companies" not in wb.sheetnames:
        return 0
    ws = wb["Companies"]
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = _clean(row[0])
        if not name or name == TRACKER_EXAMPLE_NAME:
            continue
        data = {
            "company_name": name,
            "sector": map_sector(row[1]),
            "province_city": _clean(row[2]),
            "address": _clean(row[3]),
            "owner_contact_name": _clean(row[4]),
            "phone": _clean(row[5]),
            "has_website": _clean(row[6]),
            "has_social_only": _clean(row[7]),
            "estimated_size": map_size(row[8]),
            "source": map_source(row[9]),
            "current_stage": _clean(row[10]) or "Identified",
            "last_contact_date": _iso_date(row[11]),
            "next_follow_up_date": _iso_date(row[12]),
            "notes": _clean(row[13]),
            "deal_value_dop": row[14] if isinstance(row[14], (int, float)) else None,
            "outcome_reason": _clean(row[15]),
        }
        db.insert_company(conn, data)
        count += 1
    return count


def _iso_date(v):
    if v is None:
        return None
    try:
        return v.date().isoformat()  # datetime
    except AttributeError:
        pass
    s = str(v).strip()
    return s[:10] if s else None


def main():
    reset = "--reset" in sys.argv
    db.init_db()
    where = "Postgres" if not db.IS_SQLITE else "SQLite (crm.db)"
    with db.engine.begin() as conn:
        existing = db.count_companies(conn)
        if existing and not reset:
            print(f"⚠  Database ({where}) already has {existing} companies. "
                  f"Nothing imported.\n   Run with --reset to wipe and reseed.")
            return
        if reset and existing:
            conn.execute(db.companies.delete())
            print(f"Reset: cleared {existing} existing companies.")

        n_cand = import_candidates(conn)
        n_track = import_tracker(conn)
        total = db.count_companies(conn)

    print("─" * 48)
    print(f"✔  Seed complete → {where}")
    print(f"   {n_cand} from research/san_cristobal_candidates.xlsx")
    print(f"   {n_track} from tracker/DR_Outreach_Tracker.xlsx (example row skipped)")
    print(f"   {total} companies now in the CRM.")
    print("─" * 48)


if __name__ == "__main__":
    main()
